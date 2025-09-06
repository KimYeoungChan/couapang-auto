from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
import csv
import json
import pandas as pd
from datetime import datetime
from urllib.parse import parse_qs, urlparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class CoupangPartnersWebAutomation:
    def __init__(self, headless=False):
        """
        쿠팡 파트너스 웹사이트 자동화 클래스
        
        Args:
            headless (bool): 브라우저를 숨김 모드로 실행할지 여부
        """
        self.setup_driver(headless)
        self.base_url = "https://partners.coupang.com"
        self.target_url = "https://partners.coupang.com/#affiliate/ws"
        self.products_data = []

    def setup_driver(self, headless=False):
        """Chrome 드라이버 설정"""
        chrome_options = Options()
        
        if headless:
            chrome_options.add_argument('--headless')
        
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_argument('--window-size=1920,1080') 
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        self.driver = webdriver.Chrome(options=chrome_options)
        self.wait = WebDriverWait(self.driver, 15)
        
        print("Chrome 드라이버 초기화 완료")

    def login(self, email, password):
        """
        쿠팡 파트너스에 로그인
        
        Args:
            email (str): 로그인 이메일
            password (str): 로그인 비밀번호
        """
        try:
            print("쿠팡 파트너스 로그인 중...")
            
            # 쿠팡 파트너스 메인 페이지로 이동
            print("쿠팡 파트너스 메인 페이지 접속 중...")
            self.driver.get(self.base_url)
            time.sleep(3)
            
            # 로그인 버튼 찾기 및 클릭
            print("로그인 버튼 찾는 중...")
            login_button_selectors = [
                ".ant-btn.btn-link",
                "button.ant-btn.btn-link",
                ".header-toolbar .ant-btn.btn-link",
                ".login-signup .ant-btn.btn-link",
                "//button[@class='ant-btn btn-link']",
                "//button[contains(@class, 'ant-btn') and contains(@class, 'btn-link')]",
                "//button[contains(text(), '로그인')]",
                "//span[contains(text(), '로그인')]/parent::button",
                "a[href*='login']",
                "button[class*='login']"
            ]

            login_button = None
            for selector in login_button_selectors:
                try:
                    if selector.startswith("//"):
                        elements = self.driver.find_elements(By.XPATH, selector)
                    else:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            text = element.text.strip().lower()
                            class_name = element.get_attribute('class') or ""
                            href = element.get_attribute('href') or ""
                            
                            if (
                                ('ant-btn' in class_name and 'btn-link' in class_name) or
                                any(keyword in text for keyword in ['로그인', 'login', '시작']) or 
                                'login' in href.lower()
                            ):
                                login_button = element
                                print(f"로그인 버튼 찾음: {selector}")
                                break
                    
                    if login_button:
                        break
                        
                except Exception:
                    continue
            
            if not login_button:
                print("로그인 버튼을 찾을 수 없습니다.")
                return False
            
            # 로그인 버튼 클릭
            print("로그인 버튼 클릭 중...")
            self.driver.execute_script("arguments[0].scrollIntoView(true);", login_button)
            time.sleep(1)
            
            try:
                login_button.click()
            except:
                try:
                    self.driver.execute_script("arguments[0].click();", login_button)
                except:
                    ActionChains(self.driver).move_to_element(login_button).click().perform()
            
            time.sleep(3)
            
            print("로그인 폼에 정보 입력 중...")
            
            # 이메일 입력
            email_selectors = [
                "input[type='email']",
                "input[name='email']",
                "input[name='username']", 
                "input[name='loginId']",
                "input[placeholder*='이메일']",
                "input[placeholder*='아이디']",
                "#email",
                "#loginId",
                "#username"
            ]
            
            email_input = None
            for selector in email_selectors:
                try:
                    email_input = WebDriverWait(self.driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                    if email_input and email_input.is_displayed():
                        print(f"이메일 입력창 찾음: {selector}")
                        break
                except:
                    continue
            
            if not email_input:
                print("이메일 입력창을 찾을 수 없습니다.")
                return False
            
            email_input.clear()
            email_input.send_keys(email)
            print("이메일 입력 완료")
            
            # 비밀번호 입력
            password_selectors = [
                "input[type='password']",
                "input[name='password']",
                "input[placeholder*='비밀번호']",
                "#password"
            ]
            
            password_input = None
            for selector in password_selectors:
                try:
                    password_input = self.driver.find_element(By.CSS_SELECTOR, selector)
                    if password_input and password_input.is_displayed():
                        print(f"비밀번호 입력창 찾음: {selector}")
                        break
                except:
                    continue
            
            if not password_input:
                print("비밀번호 입력창을 찾을 수 없습니다.")
                return False
            
            password_input.clear()
            password_input.send_keys(password)
            print("비밀번호 입력 완료")
            
            # 로그인 제출
            submit_selectors = [
                "button[type='submit']",
                "input[type='submit']",
                "//button[contains(text(), '로그인')]",
                "//input[@value='로그인']"
            ]
            
            submit_button = None
            for selector in submit_selectors:
                try:
                    if selector.startswith("//"):
                        submit_button = self.driver.find_element(By.XPATH, selector)
                    else:
                        submit_button = self.driver.find_element(By.CSS_SELECTOR, selector)
                    
                    if submit_button and submit_button.is_displayed():
                        print(f"로그인 제출 버튼 찾음: {selector}")
                        break
                except:
                    continue
            
            if submit_button:
                submit_button.click()
                print("로그인 제출 버튼 클릭 완료")
            else:
                password_input.send_keys(Keys.ENTER)
                print("Enter 키로 로그인 시도")
            
            # 로그인 완료 대기
            print("로그인 완료 대기 중...")
            try:
                WebDriverWait(self.driver, 15).until(
                    lambda driver: 
                    ("dashboard" in driver.current_url.lower() or
                     "main" in driver.current_url.lower() or
                     "home" in driver.current_url.lower() or
                     ("partners.coupang.com" in driver.current_url and "login" not in driver.current_url.lower()))
                )
                print("로그인 성공!")
                return True
            except:
                print("로그인 상태 확인 중...")
                time.sleep(5)
                return True
                
        except Exception as e:
            print(f"로그인 실패: {e}")
            return False

    def navigate_to_affiliate_ws(self):
        """제휴 워크스페이스로 이동"""
        try:
            print("제휴 워크스페이스로 이동 중...")
            self.driver.get(self.target_url)
            time.sleep(5)
            
            print(f"현재 URL: {self.driver.current_url}")
            return True
            
        except Exception as e:
            print(f"워크스페이스 이동 실패: {e}")
            return False

    def search_products_and_get_short_urls(self, keyword, count=10):
        """
        키워드로 상품 검색 후 첫 번째 상품에 마우스 호버하여 단축 URL 생성 및 상품 정보 저장
        
        Args:
            keyword (str): 검색할 키워드
        
        Returns:
            list: 상품 정보와 단축 URL이 포함된 딕셔너리 리스트
        """
        try:
            print(f"'{keyword}' 키워드로 상품 검색 중...")
            
            # 검색 입력창 찾기
            print("검색 입력창 찾는 중...")
            search_input_selectors = [
                ".ant-input.ant-input-lg",
                "input.ant-input.ant-input-lg",
                ".ant-input-lg",
                "input[class*='ant-input'][class*='lg']",
                "input[placeholder*='상품']",
                "input[placeholder*='검색']",
                "input[placeholder*='키워드']",
                ".search-input",
                "#search-input"
            ]
            
            search_input = None
            for selector in search_input_selectors:
                try:
                    search_input = WebDriverWait(self.driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                    if search_input and search_input.is_displayed():
                        print(f"검색 입력창 찾음: {selector}")
                        break
                except:
                    continue
            
            if not search_input:
                print("검색 입력창을 찾을 수 없습니다.")
                return None
            
            # 검색어 입력
            print(f"검색어 '{keyword}' 입력 중...")
            search_input.clear()
            search_input.send_keys(keyword)
            time.sleep(1)
            
            # 검색 버튼 클릭
            print("검색 버튼 찾는 중...")
            search_button_selectors = [
                ".search-button",
                "button.search-button", 
                "[class*='search-button']",
                ".ant-btn[class*='search']",
                "button[class*='search']",
                ".search-btn",
                ".btn-search",
                "//button[contains(@class, 'search-button')]",
                "//button[contains(text(), '검색')]"
            ]
            
            search_button = None
            for selector in search_button_selectors:
                try:
                    if selector.startswith("//"):
                        search_button = self.driver.find_element(By.XPATH, selector)
                    else:
                        search_button = self.driver.find_element(By.CSS_SELECTOR, selector)
                    
                    if search_button and search_button.is_displayed() and search_button.is_enabled():
                        print(f"검색 버튼 찾음: {selector}")
                        break
                except:
                    continue
            
            if search_button:
                search_button.click()
                print("검색 버튼 클릭 완료")
            else:
                search_input.send_keys(Keys.ENTER)
                print("Enter 키로 검색 시도")
            
            # 검색 결과 로딩 대기
            print("검색 결과 로딩 대기 중...")
            time.sleep(5)
            
            # 여러 상품 정보 추출
            products = self.extract_multiple_products_info(count)
            
            if products:
                print(f"\n총 {len(products)}개 상품 정보 추출 완료")
                
                # 각 상품에 대해 단축 URL 생성
                for i, product in enumerate(products, 1):
                    print(f"\n[{i}/{len(products)}] {product['name'][:30]}... 단축 URL 생성 중...")
                    
                    # 검색 페이지로 돌아가기
                    if i > 1:
                        self.driver.get(self.target_url)
                        time.sleep(2)
                        
                        # 다시 검색
                        search_input = self.driver.find_element(By.CSS_SELECTOR, ".ant-input.ant-input-lg")
                        search_input.clear()
                        search_input.send_keys(keyword)
                        search_input.send_keys(Keys.ENTER)
                        time.sleep(3)
                    
                    # 해당 상품에 마우스 호버하여 단축 URL 생성
                    short_url = self.get_short_url_from_hover_by_index(i-1)
                    
                    if short_url:
                        product['short_url'] = short_url
                        product['deep_link'] = ''  # 딥링크는 필요시 추가
                        print(f"[O] 단축 URL 생성 성공: {short_url}")
                    else:
                        product['short_url'] = '생성 실패'
                        product['deep_link'] = ''
                        print(f"[X] 단축 URL 생성 실패")
                
                # products_data에 추가
                self.products_data = products
                
                return products
            else:
                print("상품 정보를 추출할 수 없습니다.")
                return None
            
        except Exception as e:
            print(f"상품 검색 실패: {e}")
            return None

    def extract_multiple_products_info(self, count=10):
        """검색 결과에서 여러 상품 정보 추출"""
        try:
            print(f"{count}개 상품 정보 추출 중...")
            
            # 상품 목록 컨테이너 찾기
            product_selectors = [
                "[data-testid='product-item']",
                ".product-item", 
                ".search-product",
                ".ant-card",
                ".product-card",
                ".search-result-item",
                "[class*='product']",
                "[class*='item']",
                "[class*='card']",
                ".ant-list-item",
                ".list-item",
                "[class*='result']",
                ".ant-row .ant-col"
            ]
            
            product_elements = []
            for selector in product_selectors:
                try:
                    elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if elements and len(elements) >= 1:
                        valid_elements = []
                        for element in elements:
                            if element.is_displayed():
                                text = element.text.strip()
                                links = element.find_elements(By.TAG_NAME, "a")
                                images = element.find_elements(By.TAG_NAME, "img")
                                
                                if (text and len(text) > 10) or links or images:
                                    valid_elements.append(element)
                        
                        if len(valid_elements) >= 1:
                            product_elements = valid_elements
                            print(f"상품 목록 컨테이너 찾음: {selector} ({len(valid_elements)}개)")
                            break
                except:
                    continue
            
            if not product_elements:
                print("상품 목록을 찾을 수 없습니다.")
                return None
            
            # 여러 상품 정보 추출
            products = []
            for i in range(min(count, len(product_elements))):
                product_info = self.extract_single_product_info(product_elements[i], i+1)
                if product_info:
                    products.append(product_info)
            
            return products
            
        except Exception as e:
            print(f"상품 정보 추출 실패: {e}")
            return None
    
    def get_short_url_from_hover_by_index(self, index):
        """특정 인덱스의 상품에 마우스 호버하여 단축 URL 생성"""
        try:
            print(f"{index+1}번째 상품에 마우스 호버 중...")
            
            # 상품 목록 다시 찾기
            product_selectors = [
                "[data-testid='product-item']",
                ".product-item", 
                ".search-product",
                ".ant-card",
                ".product-card",
                ".search-result-item",
                "[class*='product']",
                "[class*='item']",
                "[class*='card']",
                ".ant-list-item",
                ".list-item",
                "[class*='result']",
                ".ant-row .ant-col"
            ]
            
            product_elements = []
            for selector in product_selectors:
                try:
                    elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if elements and len(elements) > index:
                        valid_elements = []
                        for element in elements:
                            if element.is_displayed():
                                text = element.text.strip()
                                links = element.find_elements(By.TAG_NAME, "a")
                                images = element.find_elements(By.TAG_NAME, "img")
                                
                                if (text and len(text) > 10) or links or images:
                                    valid_elements.append(element)
                        
                        if len(valid_elements) > index:
                            product_elements = valid_elements
                            break
                except:
                    continue
            
            if not product_elements or len(product_elements) <= index:
                print("상품을 찾을 수 없습니다.")
                return None
            
            # 해당 인덱스의 상품에 마우스 호버
            target_element = product_elements[index]
            
            actions = ActionChains(self.driver)
            actions.move_to_element(target_element).perform()
            time.sleep(2)  # 호버 후 버튼이 나타나길 기다림
            
            # 링크 생성 버튼 찾기
            print("링크 생성 버튼 찾는 중...")
            link_button_selectors = [
                "button.ant-btn.hover-btn.btn-generate-link",
                ".ant-btn.hover-btn.btn-generate-link",
                "button[class*='btn-generate-link']",
                "button[class*='hover-btn']",
                "//button[contains(@class, 'hover-btn')]",
                "//button[contains(text(), '링크생성')]",
                "//button[contains(text(), '링크 생성')]"
            ]
            
            link_button = None
            for selector in link_button_selectors:
                try:
                    if selector.startswith("//"):
                        elements = self.driver.find_elements(By.XPATH, selector)
                    else:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            link_button = element
                            print(f"[O] 링크 생성 버튼 찾음")
                            break
                    
                    if link_button:
                        break
                        
                except Exception:
                    continue
            
            if not link_button:
                print("[X] 링크 생성 버튼을 찾을 수 없음")
                return None
            
            # 링크 생성 버튼 클릭
            print("링크 생성 버튼 클릭 중...")
            try:
                link_button.click()
                print("[O] 링크 생성 버튼 클릭 성공")
            except:
                try:
                    self.driver.execute_script("arguments[0].click();", link_button)
                    print("[O] JavaScript로 클릭 성공")
                except Exception as e:
                    print(f"[X] 클릭 실패: {e}")
                    return None
            
            time.sleep(3)
            
            # 링크 생성 페이지로 이동 확인
            current_url = self.driver.current_url
            print(f"현재 URL: {current_url[:100]}...")
            
            if 'linkgeneration' in current_url:
                print("[O] 링크 생성 페이지로 이동 성공")
                
                # 단축 URL 찾기
                short_url = self.extract_short_url_from_page()
                return short_url
            else:
                print("[X] 링크 생성 페이지로 이동 실패")
                return None
                    
        except Exception as e:
            print(f"단축 URL 생성 실패: {e}")
            return None
    
    def get_short_url_from_hover(self):
        """첫 번째 상품에 마우스 호버하여 단축 URL 생성"""  
        try:
            print(f"첫 번째 상품에 마우스 호버 중...")
            
            # 상품 목록 컨테이너 찾기
            product_selectors = [
                "[data-testid='product-item']",
                ".product-item", 
                ".search-product",
                ".ant-card",
                ".product-card",
                ".search-result-item",
                "[class*='product']",
                "[class*='item']",
                "[class*='card']",
                ".ant-list-item",
                ".list-item",
                "[class*='result']",
                ".ant-row .ant-col"
            ]
            
            product_elements = []
            for selector in product_selectors:
                try:
                    elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if elements and len(elements) >= 3:
                        valid_elements = []
                        for element in elements:
                            if element.is_displayed():
                                text = element.text.strip()
                                links = element.find_elements(By.TAG_NAME, "a")
                                images = element.find_elements(By.TAG_NAME, "img")
                                
                                if (text and len(text) > 10) or links or images:
                                    valid_elements.append(element)
                        
                        if len(valid_elements) >= 3:
                            product_elements = valid_elements
                            print(f"상품 목록 컨테이너 찾음: {selector} ({len(valid_elements)}개)")
                            break
                except:
                    continue
            
            if not product_elements:
                print("상품 목록을 찾을 수 없습니다.")
                return None
            
            # 첫 번째 상품에 마우스 호버
            first_element = product_elements[0]
            print(f"첫 번째 상품에 마우스 호버...")
            
            actions = ActionChains(self.driver)
            actions.move_to_element(first_element).perform()
            time.sleep(2)  # 호버 후 버튼이 나타나길 기다림
            
            # 링크 생성 버튼 찾기
            print("링크 생성 버튼 찾는 중...")
            link_button_selectors = [
                "button.ant-btn.hover-btn.btn-generate-link",
                ".ant-btn.hover-btn.btn-generate-link",
                "button[class*='btn-generate-link']",
                "button[class*='hover-btn']",
                "//button[contains(@class, 'hover-btn')]",
                "//button[contains(text(), '링크생성')]",
                "//button[contains(text(), '링크 생성')]"
            ]
            
            link_button = None
            for selector in link_button_selectors:
                try:
                    if selector.startswith("//"):
                        elements = self.driver.find_elements(By.XPATH, selector)
                    else:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            link_button = element
                            print(f"[O] 링크 생성 버튼 찾음")
                            break
                    
                    if link_button:
                        break
                        
                except Exception:
                    continue
            
            if not link_button:
                print("[X] 링크 생성 버튼을 찾을 수 없음")
                return None
            
            # 링크 생성 버튼 클릭
            print("링크 생성 버튼 클릭 중...")
            try:
                link_button.click()
                print("[O] 링크 생성 버튼 클릭 성공")
            except:
                try:
                    self.driver.execute_script("arguments[0].click();", link_button)
                    print("[O] JavaScript로 클릭 성공")
                except Exception as e:
                    print(f"[X] 클릭 실패: {e}")
                    return None
            
            time.sleep(3)
            
            # 링크 생성 페이지로 이동 확인
            current_url = self.driver.current_url
            print(f"현재 URL: {current_url[:100]}...")
            
            if 'linkgeneration' in current_url:
                print("[O] 링크 생성 페이지로 이동 성공")
                
                # 단축 URL 찾기
                short_url = self.extract_short_url_from_page()
                return short_url
            else:
                print("[X] 링크 생성 페이지로 이동 실패")
                return None
                    
        except Exception as e:
            print(f"단축 URL 생성 실패: {e}")
            return None
    
    def extract_short_url_from_page(self):
        """링크 생성 페이지에서 단축 URL 추출"""
        try:
            print("단축 URL 찾는 중...")
            
            url_input_selectors = [
                ".shorten-url-wrapper > div.unselectable-input.shorten-url-input.large",
                ".shorten-url-wrapper .unselectable-input.shorten-url-input",
            ]
            
            # input 필드에서 URL 찾기
            for selector in url_input_selectors:
                try:
                    inputs = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for input_elem in inputs:
                        if input_elem.is_displayed():
                            url_value = input_elem.get_attribute('value')
                            if url_value and 'link.coupang.com' in url_value:
                                print(f"[O] 단축 URL 찾음: {url_value}")
                                return url_value
                except:
                    continue
            
            # 단축 URL이 없다면 생성 버튼 찾기
            print("단축 URL 생성 버튼 찾는 중...")
            short_url_selectors = [
                "div.shorten-url-controls > button.ant-btn.lg.shorten-url-controls-main",
                ".shorten-url-controls .ant-btn.lg.shorten-url-controls-main",
                "button[class*='shorten-url-controls']",
                "//button[contains(@class, 'shorten-url-controls-main')]",
                "//button[contains(text(), '단축')]",
                "//button[contains(text(), '짧은')]",
                "//span[contains(text(), '단축')]/parent::button",
                "button[class*='short']",
                ".ant-btn[class*='primary']"
            ]
            
            for selector in short_url_selectors:
                try:
                    if selector.startswith("//"):
                        elements = self.driver.find_elements(By.XPATH, selector)
                    else:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            text = element.text.strip()
                            # 클래스명에 shorten-url-controls가 포함되어 있거나 텍스트에 단축이 포함된 경우
                            class_name = element.get_attribute('class') or ""
                            if ('shorten-url-controls' in class_name) or ('단축' in text) or ('짧은' in text) or ('복사' in text):
                                print(f"[O] 단축 URL 생성/복사 버튼 찾음: '{text}' (class: {class_name})")
                                element.click()
                                print(f"[O] 버튼 클릭 완료")
                                time.sleep(2)
                                
                                # 클립보드 대신 바로 페이지에서 URL 추출 시도
                                time.sleep(1)
                                
                                # 다시 URL 찾기
                                for selector2 in url_input_selectors:
                                    try:
                                        inputs = self.driver.find_elements(By.CSS_SELECTOR, selector2)
                                        for input_elem in inputs:
                                            if input_elem.is_displayed():
                                                url_value = input_elem.get_attribute('value')
                                                if url_value and 'link.coupang.com' in url_value:
                                                    print(f"[O] 단축 URL 생성 완료: {url_value}")
                                                    return url_value
                                    except:
                                        continue
                                break
                except:
                    continue
            
            print("[X] 단축 URL을 찾을 수 없음")
            return None
            
        except Exception as e:
            print(f"[X] 단축 URL 추출 오류: {e}")
            return None

    def extract_single_product_info(self, element, rank):
        """개별 상품 정보 추출"""
        try:
            print(f"      === {rank}번째 상품 정보 추출 시작 ===")
            
            # 요소의 기본 정보 출력
            try:
                element_class = element.get_attribute('class') or ""
                element_tag = element.tag_name
                element_text_preview = element.text.strip()[:100] + "..." if element.text.strip() else "텍스트 없음"
                print(f"      요소 정보: tag={element_tag}, class='{element_class}', text='{element_text_preview}'")
            except:
                print(f"      요소 기본 정보 가져오기 실패")
            
            # 상품명 추출 - 전체 텍스트에서 첫 줄 추출
            print(f"      상품명 추출 시도 중...")
            
            name = ""
            
            # 먼저 전체 텍스트에서 첫 줄을 상품명으로 가져오기
            try:
                full_text = element.text.strip()
                if full_text:
                    lines = full_text.split('\n')
                    if lines and lines[0]:
                        # 첫 번째 줄이 상품명일 가능성이 높음
                        potential_name = lines[0].strip()
                        # 가격 정보나 할인율이 아닌지 확인
                        if not any(keyword in potential_name for keyword in ['%', '원', '₩', '할인']) and len(potential_name) > 5:
                            name = potential_name
                            print(f"      [O] 텍스트 첫 줄에서 상품명 찾음: '{name[:50]}...'")
            except:
                pass
            
            # 여전히 못 찾았다면 기존 방식 시도
            if not name:
                name_selectors = [
                    ".product-name", 
                    "[data-testid='product-name']", 
                    "h3", "h4", ".title",
                    ".ant-card-meta-title",
                    "[class*='title']",
                    "[class*='name']",
                    "a[href*='product']",
                    "span",
                    "div"
                ]
                
                for selector in name_selectors:
                    try:
                        name_elements = element.find_elements(By.CSS_SELECTOR, selector)
                        for name_element in name_elements[:3]:  # 처음 3개만 확인
                            text = name_element.text.strip()
                            # 상품명으로 적절한지 확인 (너무 짧거나 가격 정보가 아닌지)
                            if text and len(text) > 10 and not any(keyword in text for keyword in ['%', '원', '₩', '할인', '쿠폰']):
                                name = text.split('\n')[0]  # 여러 줄이면 첫 줄만
                                print(f"      [O] 상품명 찾음: '{name[:50]}...' (셀렉터: {selector})")
                                break
                        if name:
                            break
                    except:
                        continue
            
            if not name:
                print(f"      [X] 상품명을 찾을 수 없음")
                # 요소의 모든 텍스트를 출력해서 디버깅
                try:
                    all_text = element.text.strip()
                    print(f"      요소 전체 텍스트: '{all_text[:200]}...'")
                except:
                    pass
                return None
            
            # 상품 URL 추출
            print(f"      상품 URL 추출 시도 중...")
            link_selectors = [
                "a[href*='product']",
                "a[href*='item']", 
                "a"
            ]
            
            product_url = ""
            found_url_selector = ""
            for selector in link_selectors:
                try:
                    link_elements = element.find_elements(By.CSS_SELECTOR, selector)
                    print(f"      셀렉터 '{selector}'로 {len(link_elements)}개 링크 발견")
                    
                    for i, link_element in enumerate(link_elements):
                        href = link_element.get_attribute('href')
                        if href:
                            print(f"        링크 {i+1}: {href[:50]}...")
                            if ('product' in href or 'item' in href or 'coupang.com' in href):
                                product_url = href
                                found_url_selector = selector
                                print(f"      [O] 상품 URL 찾음: {href[:50]}... (셀렉터: {selector})")
                                break
                        else:
                            print(f"        링크 {i+1}: href 없음")
                    
                    if product_url:
                        break
                except Exception as e:
                    print(f"      URL 셀렉터 '{selector}' 실패: {e}")
                    continue
            
            if not product_url:
                print(f"      [!] 상품 URL을 찾을 수 없음")
            
            # 가격 추출
            print(f"      가격 정보 추출 시도 중...")
            price_selectors = [
                ".price", 
                "[data-testid='price']", 
                ".product-price",
                "[class*='price']",
                ".cost", ".amount"
            ]
            
            price = "가격정보없음"
            found_price_selector = ""
            for selector in price_selectors:
                try:
                    price_elements = element.find_elements(By.CSS_SELECTOR, selector)
                    print(f"      가격 셀렉터 '{selector}'로 {len(price_elements)}개 요소 발견")
                    
                    for i, price_element in enumerate(price_elements):
                        price_text = price_element.text.strip()
                        print(f"        가격 요소 {i+1}: '{price_text}'")
                        
                        if price_text and (
                            '원' in price_text or '₩' in price_text or 
                            ',' in price_text or any(char.isdigit() for char in price_text)
                        ):
                            price = price_text
                            found_price_selector = selector
                            print(f"      [O] 가격 찾음: '{price}' (셀렉터: {selector})")
                            break
                    
                    if price != "가격정보없음":
                        break
                except Exception as e:
                    # 에러 메시지 간소화
                    error_msg = "요소를 찾을 수 없음" if "no such element" in str(e) else str(e)[:50]
                    print(f"      가격 셀렉터 '{selector}' 실패: {error_msg}")
                    continue
            
            if price == "가격정보없음":
                print(f"      [!] 가격 정보를 찾을 수 없음")
                # 요소 내의 모든 숫자가 포함된 텍스트 찾아보기
                try:
                    all_text = element.text
                    import re
                    price_patterns = re.findall(r'[\d,]+원|₩[\d,]+|[\d,]+\s*원', all_text)
                    if price_patterns:
                        print(f"      발견된 가격 패턴들: {price_patterns}")
                        price = price_patterns[0]
                        print(f"      [O] 패턴에서 가격 추출: '{price}'")
                except Exception as e:
                    print(f"      패턴 매칭 실패: {e}")
            
            # 이미지 URL 추출
            print(f"      이미지 URL 추출 시도 중...")
            image_url = ""
            try:
                img_elements = element.find_elements(By.CSS_SELECTOR, "img")
                print(f"      발견된 이미지 수: {len(img_elements)}개")
                
                for i, img_element in enumerate(img_elements):
                    src = img_element.get_attribute('src')
                    data_src = img_element.get_attribute('data-src')
                    alt = img_element.get_attribute('alt') or ""
                    
                    print(f"        이미지 {i+1}: src='{src[:50] if src else 'None'}...', data-src='{data_src[:50] if data_src else 'None'}...', alt='{alt[:30]}...'")
                    
                    if src or data_src:
                        image_url = src or data_src
                        print(f"      [O] 이미지 URL 찾음: {image_url[:50]}...")
                        break
                        
                if not image_url:
                    print(f"      [!] 이미지 URL을 찾을 수 없음")
                    
            except Exception as e:
                print(f"      이미지 추출 오류: {e}")
            
            # 결과 정리
            result = {
                'rank': rank,
                'name': name,
                'price': price,
                'product_url': product_url,
                'image_url': image_url,
                'short_url': '',  # 나중에 생성
                'deep_link': ''   # 나중에 생성
            }
            
            print(f"      === {rank}번째 상품 정보 추출 완료 ===")
            print(f"      결과: 상품명='{name[:30]}...', 가격='{price}', URL={bool(product_url)}, 이미지={bool(image_url)}")
            print()
            
            return result
            
        except Exception as e:
            print(f"      [X] {rank}번째 상품 정보 추출 중 오류: {e}")
            import traceback
            print(f"      오류 상세: {traceback.format_exc()}")
            return None

    def generate_short_urls_for_all(self):
        """상위 10개 상품에 대해 단축 URL 생성"""
        try:
            if not self.products_data:
                print("단축 URL을 생성할 상품이 없습니다.")
                return
            
            # 최대 10개 상품만 처리
            products_to_process = min(10, len(self.products_data))
            print(f"\n=== 상위 {products_to_process}개 상품의 단축 URL 생성 시작 ===")
            
            for idx, product in enumerate(self.products_data[:products_to_process], 1):
                print(f"\n[{idx}/{products_to_process}] {product['name'][:50]}...")
                
                if not product.get('product_url'):
                    product['short_url'] = 'URL 없음'
                    product['deep_link'] = 'URL 없음'
                    print("  상품 URL이 없습니다.")
                    continue
                
                try:
                    # generate_single_short_url_with_info 메서드 사용
                    product_info, short_url, deep_link = self.generate_single_short_url_with_info(product['product_url'])
                    product['short_url'] = short_url
                    product['deep_link'] = deep_link
                    
                    if short_url and short_url not in ["단축 URL 생성 실패", "처리 오류", "링크 생성 버튼 없음", "클릭 실패"]:
                        print(f"  ✓ 단축 URL 생성 성공")
                    else:
                        print(f"  ✗ 단축 URL 생성 실패: {short_url}")
                        
                except Exception as e:
                    print(f"  ✗ 오류 발생: {e}")
                    product['short_url'] = '생성 오류'
                    product['deep_link'] = '생성 오류'
                
                # 잠시 대기 (서버 부하 방지)
                if idx < products_to_process:
                    time.sleep(2)
            
            print(f"\n=== {products_to_process}개 상품 단축 URL 생성 완료 ===")
            
        except Exception as e:
            print(f"단축 URL 생성 실패: {e}")

    def generate_single_short_url_with_info(self, product_url):
        """상품 정보 추출과 단축 URL 생성을 동시에 수행"""
        try:
            print(f"\n상품 페이지로 이동: {product_url[:50]}...")
            
            # 새 탭에서 상품 페이지 열기
            self.driver.execute_script("window.open('');")
            self.driver.switch_to.window(self.driver.window_handles[-1])
            
            self.driver.get(product_url)
            time.sleep(3)
            
            # 링크 생성 버튼 클릭
            print(f"링크 생성 버튼 찾는 중...")
            link_generation_selectors = [
                "button.ant-btn.hover-btn.btn-generate-link",
                ".ant-btn.hover-btn.btn-generate-link",
                "button[class*='btn-generate-link']",
                "//button[contains(@class, 'btn-generate-link')]",
                "//button[contains(text(), '링크생성')]",
                "//button[contains(text(), '링크 생성')]"
            ]
            
            link_button = None
            for selector in link_generation_selectors:
                try:
                    if selector.startswith("//"):
                        elements = self.driver.find_elements(By.XPATH, selector)
                    else:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            text = element.text.strip()
                            class_name = element.get_attribute('class') or ""
                            
                            if ('링크' in text and '생성' in text) or 'btn-generate-link' in class_name:
                                link_button = element
                                print(f"[O] 링크 생성 버튼 찾음: '{text}'")
                                break
                    
                    if link_button:
                        break
                        
                except Exception:
                    continue
            
            if not link_button:
                print(f"[X] 링크 생성 버튼을 찾을 수 없음")
                self.driver.close()
                self.driver.switch_to.window(self.driver.window_handles[0])
                return None, "링크 생성 버튼 없음", "딥링크 생성 실패"
            
            # 링크 생성 버튼 클릭
            print(f"링크 생성 버튼 클릭 중...")
            try:
                self.driver.execute_script("arguments[0].scrollIntoView(true);", link_button)
                time.sleep(1)
                link_button.click()
                print(f"[O] 링크 생성 버튼 클릭 성공")
            except Exception:
                try:
                    self.driver.execute_script("arguments[0].click();", link_button)
                    print(f"[O] JavaScript 클릭 성공")
                except Exception as e:
                    print(f"[X] 클릭 실패: {e}")
                    self.driver.close()
                    self.driver.switch_to.window(self.driver.window_handles[0])
                    return None, "클릭 실패", "딥링크 생성 실패"
            
            time.sleep(3)
            
            # 링크 생성 페이지로 이동 확인 및 상품 정보 추출
            current_url = self.driver.current_url
            print(f"현재 URL: {current_url[:100]}...")
            
            product_info = None
            short_url = "단축 URL 생성 실패"
            deep_link = "딥링크 생성 실패"
            
            if 'linkgeneration' in current_url:
                print(f"[O] 링크 생성 페이지로 이동 성공")
                
                # URL에서 상품 정보 추출
                product_info = self.extract_product_info_from_url(current_url)
                
                # 단축 URL 찾기 (이미 생성되어 있을 수 있음)
                print(f"단축 URL 찾는 중...")
                url_input_selectors = [
                    ".shorten-url-wrapper > div.unselectable-input.shorten-url-input.large",
                    ".shorten-url-wrapper .unselectable-input.shorten-url-input",
                ]
                
                # input 필드에서 URL 찾기
                for selector in url_input_selectors:
                    try:
                        inputs = self.driver.find_elements(By.CSS_SELECTOR, selector)
                        for input_elem in inputs:
                            if input_elem.is_displayed():
                                # div 요소의 경우 textContent 사용
                                url_value = input_elem.text.strip()
                                if not url_value:
                                    url_value = input_elem.get_attribute('textContent')
                                if not url_value:
                                    url_value = input_elem.get_attribute('value')
                                    
                                if url_value and 'link.coupang.com' in url_value:
                                    short_url = url_value
                                    print(f"[O] 단축 URL 찾음: {url_value[:50]}...")
                                    break
                    except:
                        continue
                
                # 단축 URL이 없다면 생성 버튼 찾기
                if short_url == "단축 URL 생성 실패":
                    print(f"단축 URL 생성 버튼 찾는 중...")
                    short_url_selectors = [
                        "button.ant-btn.lg.shorten-url-controls-main",
                        ".ant-btn.lg.shorten-url-controls-main",
                        "button[class*='shorten-url-controls']",
                        "//button[contains(@class, 'shorten-url-controls-main')]",
                        "//button[contains(text(), '단축')]",
                        "//button[contains(text(), '짧은')]",
                        "//span[contains(text(), '단축')]/parent::button",
                        "button[class*='short']",
                        ".ant-btn[class*='primary']"
                    ]
                    
                    for selector in short_url_selectors:
                        try:
                            if selector.startswith("//"):
                                elements = self.driver.find_elements(By.XPATH, selector)
                            else:
                                elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                            
                            for element in elements:
                                if element.is_displayed() and element.is_enabled():
                                    text = element.text.strip()
                                    # 클래스명에 shorten-url-controls가 포함되어 있거나 텍스트에 단축이 포함된 경우
                                    class_name = element.get_attribute('class') or ""
                                    if ('shorten-url-controls' in class_name) or ('단축' in text) or ('짧은' in text) or ('복사' in text):
                                        print(f"[O] 단축 URL 생성/복사 버튼 찾음: '{text}' (class: {class_name})")
                                        
                                        # 클립보드 초기화 (이전 내용 제거)
                                        try:
                                            import pyperclip
                                            pyperclip.copy("")  # 클립보드 비우기
                                        except:
                                            pass
                                        
                                        element.click()
                                        print(f"[O] 버튼 클릭 완료")
                                        time.sleep(3)
                                        
                                        # 클립보드에서 URL 가져오기
                                        try:
                                            import pyperclip
                                            clipboard_content = pyperclip.paste()
                                            if clipboard_content and 'link.coupang.com' in clipboard_content:
                                                short_url = clipboard_content
                                                print(f"[O] 클립보드에서 단축 URL 복사됨: {short_url[:50]}...")
                                        except:
                                            pass
                                        
                                        # 다시 URL 찾기
                                        for selector2 in url_input_selectors:
                                            try:
                                                inputs = self.driver.find_elements(By.CSS_SELECTOR, selector2)
                                                for input_elem in inputs:
                                                    if input_elem.is_displayed():
                                                        url_value = input_elem.get_attribute('value')
                                                        if url_value and 'link.coupang.com' in url_value:
                                                            short_url = url_value
                                                            print(f"[O] 단축 URL 생성 완료")
                                                            break
                                                if short_url != "단축 URL 생성 실패":
                                                    break
                                            except:
                                                continue
                                        break
                            if short_url != "단축 URL 생성 실패":
                                break
                        except:
                            continue
                
                # 여전히 못 찾았다면 페이지 소스에서 정규식으로 찾기
                if short_url == "단축 URL 생성 실패":
                    try:
                        page_text = self.driver.page_source
                        import re
                        short_urls = re.findall(r'https://link\.coupang\.com/[^\s"<>]+', page_text)
                        if short_urls:
                            short_url = short_urls[0]
                            print(f"[O] 페이지 소스에서 단축 URL 찾음")
                    except:
                        pass
                
            else:
                print(f"[X] 링크 생성 페이지로 이동 실패")
            
            # 탭 닫기
            self.driver.close()
            self.driver.switch_to.window(self.driver.window_handles[0])
            
            return product_info, short_url, deep_link
            
        except Exception as e:
            print(f"[X] 처리 중 오류: {e}")
            if len(self.driver.window_handles) > 1:
                self.driver.close()
                self.driver.switch_to.window(self.driver.window_handles[0])
            return None, "처리 오류", "딥링크 생성 실패"

    def save_results_to_excel(self, keyword, filename=None):
        """결과를 엑셀 파일로 저장"""
        if not self.products_data:
            print("저장할 데이터가 없습니다.")
            return None
        
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"쿠팡파트너스_TOP{len(self.products_data)}_{keyword}_{timestamp}.xlsx"
            
            # 데이터 포맷팅
            formatted_data = []
            for product in self.products_data:
                formatted_data.append({
                    # '순위': product['rank'],
                    '상품명': product['name'],
                    '가격': product['price'],
                    '원본URL': product['product_url'],
                    '단축URL': product['short_url'],
                    # '딥링크': product['deep_link'],
                    '이미지URL': product['image_url'],
                    '생성일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    '단축URL상태': '성공' if product['short_url'] != '단축 URL 생성 실패' else '실패'
                })
            
            df = pd.DataFrame(formatted_data)
            
            # 엑셀 파일 저장 (스타일 적용)
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='TOP상품분석', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['TOP상품분석']
                
                # 스타일 적용
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                success_fill = PatternFill(start_color="D4F4DD", end_color="D4F4DD", fill_type="solid")
                fail_fill = PatternFill(start_color="F4D4D4", end_color="F4D4D4", fill_type="solid")
                
                cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # 헤더 스타일
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # 데이터 셀 스타일
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                             min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.alignment = cell_alignment
                        cell.border = border
                        
                        # 상태에 따른 색상
                        if cell.column_letter == 'I':  # 단축URL상태 컬럼
                            if cell.value == '성공':
                                cell.fill = success_fill
                            elif cell.value == '실패':
                                cell.fill = fail_fill
                
                # 컬럼 너비 조정
                column_widths = {
                    'A': 8,   'B': 40,  'C': 15,  'D': 50,  'E': 50,
                    'F': 50,  'G': 30,  'H': 20,  'I': 12
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 행 높이 조정
                for row in range(1, worksheet.max_row + 1):
                    worksheet.row_dimensions[row].height = 25
                worksheet.row_dimensions[1].height = 35
            
            print(f"엑셀 파일 저장 완료: {filename}")
            return filename
            
        except Exception as e:
            print(f"엑셀 파일 저장 실패: {e}")
            return None

    def print_results(self, keyword):
        """검색 결과 출력"""
        if not self.products_data:
            print("출력할 결과가 없습니다.")
            return
        
        print(f"\n=== '{keyword}' TOP {len(self.products_data)} 검색 결과 ===")
        print("=" * 80)
        
        success_count = 0
        for product in self.products_data:
            status = "성공" if product['short_url'] != '단축 URL 생성 실패' else "실패"
            if status == "성공":
                success_count += 1
            
            print(f"{product['rank']}. {product['name']}")
            print(f"   가격: {product['price']}")
            print(f"   단축URL: {product['short_url']}")
            print(f"   상태: {status}")
            print()
        
        print(f"총 {len(self.products_data)}개 상품 중 {success_count}개 단축URL 생성 성공")

    def close(self):
        """브라우저 종료"""
        if hasattr(self, 'driver'):
            self.driver.quit()
            print("브라우저 종료")

def main():
    # 설정값
    EMAIL = "alstmd9708@naver.com"  # 실제 이메일로 변경
    PASSWORD = "tmddlsp1!"  # 실제 비밀번호로 변경
    KEYWORD = "미니선풍기"        # 검색할 키워드
    TOP_COUNT = 10                  # 추출할 상품 개수
    
    automation = None
    
    try:
        print("=" * 60)
        print("쿠팡 파트너스 상위 10개 상품 단축 URL 생성 및 엑셀 저장")
        print("=" * 60)
        print(f"검색 키워드: {KEYWORD}")
        print(f"추출 개수: TOP {TOP_COUNT}개 상품")
        print("=" * 60)
        
        # 자동화 클래스 초기화
        automation = CoupangPartnersWebAutomation(headless=False)
        
        # 로그인
        if not automation.login(EMAIL, PASSWORD):
            print("로그인에 실패했습니다. 수동으로 로그인을 완료한 후 계속하세요.")
            input("로그인을 완료했으면 Enter를 눌러주세요...")
        
        # 제휴 워크스페이스로 이동
        if not automation.navigate_to_affiliate_ws():
            print("워크스페이스 이동에 실패했습니다.")
            return
        
        # 여러 상품 검색 및 단축 URL 생성
        products = automation.search_products_and_get_short_urls(KEYWORD, TOP_COUNT)
        
        if products:
            print(f"\n=== 단축 URL 생성 완료! ===")
            success_count = sum(1 for p in products if p.get('short_url') and p['short_url'] != '생성 실패')
            print(f"\n총 {len(products)}개 상품 중 {success_count}개 성공")
            
            for i, product in enumerate(products, 1):
                print(f"\n{i}. {product['name'][:50]}...")
                print(f"   가격: {product['price']}")
                print(f"   단축 URL: {product.get('short_url', '없음')}")
            
            # 엑셀 파일로 저장
            filename = automation.save_results_to_excel(KEYWORD)
            if filename:
                print(f"\n엑셀 파일 저장 완료: {filename}")
        else:
            print("상품 정보 추출에 실패했습니다.")
        
    except KeyboardInterrupt:
        print("\n사용자에 의해 중단되었습니다.")
    except Exception as e:
        print(f"예상치 못한 오류: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if automation:
            automation.close()

if __name__ == "__main__":
    print("필요한 라이브러리:")
    print("  pip install selenium pandas openpyxl")
    print("=" * 60)
    main()